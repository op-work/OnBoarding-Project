"""
Import Service Component
Parses CSV, Excel, and JSON files, normalizes 71 HR system columns, tolerates missing columns, and ingests employee data into the database (supports UPSERT/updating existing records).
"""

import io
import csv
import json
import datetime
import openpyxl
from sqlalchemy.orm import Session
from models import Associate
from services.associate_service import AssociateService
from utils.logger import app_logger

# Exact user-specified HR system columns (71 columns)
ALL_HR_COLUMNS = [
    "Employee Number", "First Name", "Middle Name", "Last Name", "Display Name", "Full Name",
    "Work Email", "Date Of Birth", "Gender", "Marital Status", "Marriage Date", "Blood Group",
    "Physically Handicapped", "Nationality", "Mobile Phone", "Work Phone", "Home Phone",
    "Personal Email", "Current Address Line 1", "Current Address Line 2", "Current Address City",
    "Current Address State", "Current Address Zip", "Current Address Country",
    "Permanent Address Line 1", "Permanent Address Line 2", "Permanent Address City",
    "Permanent Address State", "Permanent Address Zip", "Permanent Address Country",
    "Father Name", "Mother Name", "Spouse Name", "Children Names", "Attendance Number",
    "Location", "Location Country", "Legal Entity", "Business Unit", "Department",
    "Sub Department", "Job Title", "Secondary Job Title", "Reporting To",
    "Reporting Manager Employee Number", "Dotted Line Manager", "Date Joined", "Leave Plan",
    "Band", "Pay Grade", "Time Type", "Worker Type", "Shift Policy Name",
    "Weekly Off Policy Name", "Attendance Time Tracking Policy", "Attendance Capture Scheme",
    "Holiday List Name", "Expense Policy Name", "Notice Period", "PAN Number", "Aadhaar Number",
    "PF Number", "UAN Number", "Employment Status", "Exit Date", "Comments", "Exit Status",
    "Termination Type", "Termination Reason", "Resignation Note", "Cost Center"
]


def normalize_key(key: str) -> str:
    """Normalizes string keys for case-insensitive and whitespace-flexible matching."""
    if not key:
        return ""
    return str(key).strip().lower().replace("_", " ").replace("-", " ")


class ImportService:
    """Handles multi-format parsing, flexible column mapping, missing column tolerance, and ingestion with UPSERT support."""

    @staticmethod
    def detect_file_format(file_bytes: bytes, file_name: str) -> tuple[bool, str, str]:
        """
        Detects and validates file format from file extension and signature bytes.
        Returns: (is_supported: bool, format_display_name: str, extension: str)
        """
        if not file_name or not file_bytes:
            return False, "Unsupported File", ""

        name_lower = file_name.lower()
        header_bytes = file_bytes[:10]

        # 1. Check Excel (.xlsx / .xls)
        if name_lower.endswith(".xlsx") or header_bytes.startswith(b"PK\x03\x04"):
            return True, "Excel Spreadsheet (.xlsx)", ".xlsx"
        if name_lower.endswith(".xls") or header_bytes.startswith(b"\xd0\xcf\x11\xe0"):
            return True, "Legacy Excel (.xls)", ".xls"

        # 2. Check JSON (.json)
        if name_lower.endswith(".json"):
            return True, "JSON Dataset (.json)", ".json"

        # 3. Check CSV (.csv)
        if name_lower.endswith(".csv"):
            return True, "CSV Delimited (.csv)", ".csv"

        return False, "File Not Supported", ""

    @staticmethod
    def parse_file(file_bytes: bytes, file_name: str) -> list[dict]:
        """
        Parses raw file bytes (CSV, Excel .xlsx, or JSON) into a list of raw row dictionaries.
        """
        file_name_lower = file_name.lower()
        rows = []

        if file_name_lower.endswith(".csv"):
            content = file_bytes.decode("utf-8-sig", errors="ignore")
            reader = csv.DictReader(io.StringIO(content))
            for r in reader:
                rows.append({k: (v.strip() if isinstance(v, str) else v) for k, v in r.items() if k})

        elif file_name_lower.endswith(".xlsx") or file_name_lower.endswith(".xls"):
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active
            header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = {}
                for idx, val in enumerate(row):
                    if idx < len(header_row) and header_row[idx]:
                        row_dict[header_row[idx]] = str(val).strip() if val is not None else ""
                if row_dict:
                    rows.append(row_dict)

        elif file_name_lower.endswith(".json"):
            content = file_bytes.decode("utf-8", errors="ignore")
            data = json.loads(content)
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict):
                        rows.append({str(k): (str(v).strip() if v is not None else "") for k, v in item.items()})
            elif isinstance(data, dict):
                rows.append({str(k): (str(v).strip() if v is not None else "") for k, v in data.items()})

        return rows

    @staticmethod
    def parse_date(date_str: str) -> datetime.date | None:
        """Parses multiple date string formats into datetime.date object."""
        if not date_str:
            return None
        date_str = str(date_str).strip()
        formats = [
            "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d",
            "%d %b %Y", "%d %B %Y", "%Y-%m-%d %H:%M:%S"
        ]
        for fmt in formats:
            try:
                return datetime.datetime.strptime(date_str.split(" ")[0], fmt).date()
            except Exception:
                pass
        return None

    @staticmethod
    def map_row_to_associate_schema(row_dict: dict) -> dict:
        """
        Flexible Column Mapping Logic:
        Maps raw uploaded headers to system fields. Missing columns remain empty without breaking execution.
        """
        norm_row = {normalize_key(k): str(v).strip() for k, v in row_dict.items() if k}

        def get_val(possible_headers: list[str]) -> str:
            for h in possible_headers:
                n_h = normalize_key(h)
                if n_h in norm_row and norm_row[n_h]:
                    return norm_row[n_h]
            return ""

        emp_number = get_val(["Employee Number", "Employee ID", "Emp ID", "Emp Code"])
        first_name = get_val(["First Name"])
        middle_name = get_val(["Middle Name"])
        last_name = get_val(["Last Name"])
        display_name = get_val(["Display Name"])
        full_name_raw = get_val(["Full Name", "Name", "Associate Name"])

        if display_name and not display_name.replace(" ", "").replace("-", "").isdigit():
            full_name = display_name
        elif full_name_raw:
            full_name = full_name_raw
        elif first_name or last_name:
            full_name = " ".join(filter(None, [first_name, middle_name, last_name])).strip()
        else:
            full_name = emp_number or "New Employee"

        work_email = get_val(["Work Email", "Official Email", "Email ID"])
        personal_email = get_val(["Personal Email", "Personal Email ID", "Email"])
        primary_email = work_email or personal_email or f"emp.{emp_number or 'temp'}@company.com"

        date_joined_raw = get_val(["Date Joined", "DOJ", "Date of Joining", "Joining Date"])
        date_of_joining = ImportService.parse_date(date_joined_raw) or datetime.date.today()

        job_title = get_val(["Job Title", "Designation", "Secondary Job Title", "Role"]) or "Associate"
        department = get_val(["Department", "Business Unit", "Sub Department", "Legal Entity"]) or "Engineering"
        location = get_val(["Location", "City", "Location Country", "Current Address City"]) or "Bangalore"
        
        worker_type = get_val(["Worker Type", "Time Type", "Work Mode"])
        if "virt" in worker_type.lower() or "remot" in worker_type.lower() or "home" in worker_type.lower():
            work_mode = "Virtual"
        else:
            work_mode = "In-Person"

        reporting_manager = get_val(["Reporting To", "Reporting Manager Employee Number", "Dotted Line Manager"]) or "HR Manager"
        name_as_per_aadhar_raw = get_val(["Name as per Aadhar", "Aadhar Name", "Name on Aadhaar", "Name As Per Aadhaar"])
        if name_as_per_aadhar_raw and not name_as_per_aadhar_raw.replace(" ", "").replace("-", "").isdigit():
            name_as_per_aadhar = name_as_per_aadhar_raw
        else:
            name_as_per_aadhar = full_name
        
        exit_date_raw = get_val(["Exit Date", "Last Working Day", "LWD"])
        last_working_day = ImportService.parse_date(exit_date_raw)

        mobile_phone = get_val(["Mobile Phone", "Work Phone", "Home Phone"]) or "9876543210"

        c_addr1 = get_val(["Current Address Line 1"])
        c_addr2 = get_val(["Current Address Line 2"])
        c_city = get_val(["Current Address City"])
        c_state = get_val(["Current Address State"])
        c_zip = get_val(["Current Address Zip"])
        shipment_address = ", ".join(filter(None, [c_addr1, c_addr2, c_city, c_state, c_zip]))

        mapped_record = {
            "display_name": display_name or full_name,
            "full_name": full_name,
            "employee_id": emp_number,
            "email": primary_email,
            "personal_email": personal_email or primary_email,
            "designation": job_title,
            "department": department,
            "location": location,
            "date_of_joining": date_of_joining,
            "work_mode": work_mode,
            "reporting_manager": reporting_manager,
            "name_as_per_aadhar": name_as_per_aadhar,
            "is_fresher": False if last_working_day else True,
            "last_working_day": last_working_day,
            "phone": mobile_phone,
            "asset_shipment_address": shipment_address,
            "raw_data": row_dict
        }
        return mapped_record

    @staticmethod
    def process_and_validate(rows: list[dict], db: Session) -> tuple[list[dict], list[dict]]:
        """
        Validates parsed rows against database. Supports updating existing employee profiles (UPSERT).
        """
        valid_records = []
        invalid_records = []

        existing_emails = {}
        existing_emp_ids = {}

        for a in db.query(Associate).all():
            if a.personal_email:
                existing_emails[a.personal_email.lower()] = a
            if a.work_email:
                existing_emails[a.work_email.lower()] = a
            if a.employee_id:
                existing_emp_ids[a.employee_id.lower()] = a

        seen_emails_in_batch = set()

        for idx, r in enumerate(rows, start=1):
            mapped = ImportService.map_row_to_associate_schema(r)
            email_lower = mapped["email"].lower()
            emp_id_lower = mapped["employee_id"].lower() if mapped["employee_id"] else ""

            status = "Create"
            issue = ""

            if not mapped["full_name"]:
                status = "Error"
                issue = "Full Name missing"
            elif email_lower in seen_emails_in_batch:
                status = "Duplicate"
                issue = f"Duplicate email {mapped['email']} repeated within uploaded file"
            elif emp_id_lower and emp_id_lower in existing_emp_ids:
                status = "Update"
                issue = f"Will update existing employee ({mapped['employee_id']})"
            elif email_lower in existing_emails:
                status = "Update"
                issue = f"Will update existing employee with email ({mapped['email']})"

            mapped["row_index"] = idx
            mapped["import_status"] = status
            mapped["import_issue"] = issue

            if status in ["Create", "Update"]:
                valid_records.append(mapped)
                seen_emails_in_batch.add(email_lower)
            else:
                invalid_records.append(mapped)

        return valid_records, invalid_records

    @staticmethod
    def ingest_records(db: Session, records: list[dict]) -> list[Associate]:
        """Bulk ingests or updates (UPSERT) validated records into the SQLite database."""
        processed_associates = []

        for rec in records:
            emp_id = rec.get("employee_id")
            email = rec.get("email")

            # Check if associate exists for update
            existing = None
            if emp_id:
                existing = db.query(Associate).filter(Associate.employee_id == emp_id).first()
            if not existing and email:
                existing = db.query(Associate).filter((Associate.personal_email == email) | (Associate.work_email == email)).first()

            if existing:
                # Update existing associate fields
                if rec.get("display_name"):
                    existing.display_name = rec["display_name"]
                if rec.get("full_name"):
                    full = rec["full_name"].strip()
                    if full.lower().startswith("associate ") and len(full.split()) > 1:
                        full = " ".join(full.split()[1:])
                    names = full.split(" ")
                    existing.first_name = names[0]
                    existing.last_name = " ".join(names[1:]) if len(names) > 1 else ""
                if rec.get("designation"):
                    existing.designation = rec["designation"]
                if rec.get("department"):
                    existing.department = rec["department"]
                if rec.get("location"):
                    existing.location = rec["location"]
                if rec.get("work_mode"):
                    existing.work_mode = rec["work_mode"]
                if rec.get("reporting_manager"):
                    existing.reporting_manager = rec["reporting_manager"]
                if rec.get("date_of_joining"):
                    existing.date_of_joining = rec["date_of_joining"]
                if rec.get("phone"):
                    existing.phone = rec["phone"]
                if rec.get("asset_shipment_address"):
                    existing.asset_shipment_address = rec["asset_shipment_address"]
                if rec.get("name_as_per_aadhar"):
                    existing.name_as_per_aadhar = rec["name_as_per_aadhar"]

                db.commit()
                db.refresh(existing)
                processed_associates.append(existing)
            else:
                # Create new associate
                full = (rec.get("full_name") or "").strip()
                if full.lower().startswith("associate ") and len(full.split()) > 1:
                    full = " ".join(full.split()[1:])
                names = full.split(" ") if full else [""]
                first_n = names[0]
                last_n = " ".join(names[1:]) if len(names) > 1 else ""

                assoc_data = {
                    "display_name": rec.get("display_name") or full,
                    "name_as_per_aadhar": rec["name_as_per_aadhar"],
                    "first_name": first_n,
                    "last_name": last_n,
                    "personal_email": rec["personal_email"],
                    "date_of_joining": rec["date_of_joining"],
                    "is_fresher": rec["is_fresher"],
                    "last_working_day": rec["last_working_day"],
                    "designation": rec["designation"],
                    "location": rec["location"],
                    "work_mode": rec["work_mode"],
                    "asset_shipment_address": rec["asset_shipment_address"],
                    "department": rec["department"],
                    "phone": rec["phone"],
                    "reporting_manager": rec["reporting_manager"]
                }
                if emp_id:
                    assoc_data["employee_id"] = emp_id

                assoc = AssociateService.create_associate(db, assoc_data, is_draft=False)
                processed_associates.append(assoc)

        app_logger.info(f"BULK IMPORT: Successfully processed/updated {len(processed_associates)} associate records.")
        return processed_associates

    @staticmethod
    def generate_sample_file(file_type: str = "csv") -> bytes:
        """Generates downloadable sample files containing all 71 exact user column headers."""
        sample_row = {col: "" for col in ALL_HR_COLUMNS}
        sample_row.update({
            "Employee Number": "EMP9001",
            "First Name": "Robert",
            "Middle Name": "Alexander",
            "Last Name": "Vance",
            "Display Name": "Robert Vance",
            "Full Name": "Robert Alexander Vance",
            "Work Email": "robert.vance@company.com",
            "Date Of Birth": "1992-05-14",
            "Gender": "Male",
            "Marital Status": "Single",
            "Blood Group": "O+",
            "Nationality": "Indian",
            "Mobile Phone": "9876543210",
            "Personal Email": "robert.vance@example.com",
            "Current Address Line 1": "Suite 402, Tech Park",
            "Current Address City": "Bangalore",
            "Current Address State": "Karnataka",
            "Current Address Zip": "560103",
            "Current Address Country": "India",
            "Location": "Bangalore",
            "Department": "Engineering",
            "Job Title": "Lead Systems Architect",
            "Reporting To": "Sarah Jenkins",
            "Date Joined": "2026-09-01",
            "Worker Type": "In-Person",
            "Aadhaar Number": "1234 5678 9012",
            "Employment Status": "Active"
        })

        if file_type.lower() == "json":
            json_str = json.dumps([sample_row], indent=2)
            return json_str.encode("utf-8")

        elif file_type.lower() in ["xlsx", "excel"]:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employee Data"
            ws.append(ALL_HR_COLUMNS)
            ws.append([sample_row[col] for col in ALL_HR_COLUMNS])
            out = io.BytesIO()
            wb.save(out)
            return out.getvalue()

        else:
            out = io.StringIO()
            writer = csv.DictWriter(out, fieldnames=ALL_HR_COLUMNS)
            writer.writeheader()
            writer.writerow(sample_row)
            return out.getvalue().encode("utf-8")
